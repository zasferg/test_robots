import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from pydantic_core._pydantic_core import ValidationError
from robots.models import Robot
from robots.validators import RobotSchema
from openpyxl import Workbook
from datetime import datetime, timedelta
from django.db.models import Count


@csrf_exempt
def create_robot(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            check_robot_in_db = Robot.objects.filter(serial=data["serial"]).exists()
            if check_robot_in_db:
                return JsonResponse(
                    {"status": "error", "message": "Robot already exists"}, status=400
                )
            robot_data = RobotSchema(**data)
            robot = Robot.objects.create(
                serial=robot_data.serial,
                model=robot_data.model,
                version=robot_data.version,
                created=robot_data.created,
            )
            if robot:
                return JsonResponse(
                    {"status": "success", "data": f"{robot}"},
                    status=201,
                )
        except json.JSONDecodeError:
            return JsonResponse(
                {"status": "error", "message": "Invalid JSON"}, status=400
            )
        except ValidationError as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
        except ValueError as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    return JsonResponse(
        {"status": "error", "message": "Invalid request method"}, status=405
    )


class HttpResponse:
    pass


from django.http import JsonResponse
from openpyxl import Workbook
from django.db.models import Count
from datetime import datetime, timedelta
from .models import Robot


def generate_report(request):
    try:
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)
        robots = Robot.objects.filter(created__range=[start_date, end_date])

        if not robots:
            return JsonResponse(
                {"status": "error", "message": "No data available for the past week"},
                status=500,
            )

        wb = Workbook()
        models = robots.values_list("model", flat=True).distinct()

        for model in models:
            ws = wb.create_sheet(title=model)
            ws.append(["Model", "Version", "Count for the week"])

            versions = (
                robots.filter(model=model)
                .values("version")
                .annotate(count=Count("version"))
            )
            if not versions:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "No data available for model: " + model,
                    },
                    status=500,
                )

            for version in versions:
                ws.append([model, version["version"], version["count"]])

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        try:
            filename = "robot_report.xlsx"
            wb.save(filename)
            return JsonResponse(
                {"status": "success", "message": f"File generated: {filename}"},
                status=200,
            )

        except Exception as e:
            return JsonResponse(
                {"status": "error", "message": "Error saving Excel file: " + str(e)},
                status=500,
            )

    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": "Unexpected error: " + str(e)}, status=500
        )
